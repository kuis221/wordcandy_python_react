from __future__ import division
from rest_framework.response import Response
from rest_framework.generics import GenericAPIView, RetrieveUpdateAPIView
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework import status
from rest_framework.authentication import TokenAuthentication, SessionAuthentication, BasicAuthentication

from django.contrib.auth import login
from django.http import HttpResponse
from django.conf import settings
from django.core.cache import cache

import math
import random
from amazon.api import AmazonAPI, AmazonSearch
import time
import requests
from datetime import datetime
import boto3
from cStringIO import StringIO
import io
import json
import urllib3
import traceback
from django.http import HttpResponseRedirect

from .serializers import *
from .models import Shop, Subscribe, Word, Product

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.writer.excel import save_virtual_workbook
from trademark import marker
import requests
import after_response

from rest_framework_tracking.mixins import LoggingMixin
from rest_framework.authentication import TokenAuthentication


from oauth2client.client import OAuth2WebServerFlow
from oauth2client.tools import run_flow
from oauth2client.file import Storage


class OAuthView(LoggingMixin, GenericAPIView):
    authentication_classes = (
        BasicAuthentication, TokenAuthentication, SessionAuthentication)

    def get(self, request):
        CLIENT_ID = '753175684403-6tp6m9jplfhp8gbm38kms7kn8fgs18d0.apps.googleusercontent.com'
        CLIENT_SECRET = 'z8dXwGeZDLXq9VMJGhZZf3NR'


        flow = OAuth2WebServerFlow(client_id=CLIENT_ID,
                                   client_secret=CLIENT_SECRET,
                                   scope='https://spreadsheets.google.com/feeds https://docs.google.com/feeds',
                                   redirect_uri='http://localhost:8000/profile/')

        auth_uri = flow.step1_get_authorize_url()

        print("access_token: {}".format(auth_uri))

        return Response({'auth': auth_uri})


def monthly_sales_estimate(bsr):
    if bsr != None:
        bsr = int(bsr)
        if 10000000000 <= bsr >= 9180:
            return 1

        if 9179 <= bsr >= 6033:
            return random.randint(2, 10)

        if 6032 <= bsr >= 3545:
            return random.randint(50, 100)

        if 3544 <= bsr >= 1714:
            return random.randint(500, 1000)

        if 1713 <= bsr >= 541:
            return random.randint(5000, 10000)

        if 540 <= bsr >= 35:
            return random.randint(50000, 100000)

        if 34 <= bsr >= 28:
            return random.randint(100000, 110000)

        if 27 <= bsr >= 7:
            return random.randint(130000, 160000)

        return random.randint(165000, 170000)
    else:
        return 0


@after_response.enable
def amazon_products(data):
    for item in data['result']:
        product = False
        #Product.objects.filter(asin=item['asin'])
        if not product:
            try:
                Product.objects.create(title=item['title'],
                                       sales_rank=item['sales_rank'],
                                       monthly_sales_estimate=item['monthly_sales_estimate'],
                                       asin=item['asin'],
                                       small_image_url=item['small_image_url'],
                                       reviews=item['reviews'],
                                       detail_page_url=item['detail_page_url'],
                                       features=item['features'],
                                       type=item['type'])

            except Exception as e:
                pass

        else:
            product.features = item['features']
            product.type = item['type']
            product.save(['features', 'type'])





class AmazonProductsView(LoggingMixin, GenericAPIView):
    authentication_classes = (
        BasicAuthentication, TokenAuthentication, SessionAuthentication)

    def getProducts(self, request, format=None):
        amazon = AmazonAPI(settings.AMAZON_ACCESS_KEY,
                           settings.AMAZON_SECRET_KEY, settings.AMAZON_ASSOC_TAG)
        data = {'result': []}
        tags = request.GET.get('tags', '')
        range = request.GET.get('range', '')

        for tag in tags.split(','):
            if range == 'low':
                print("range", range)

                products = amazon.search(
                    Keywords='{} shirt'.format(tag), SearchIndex='Apparel', BrowseNode="9056987011",
                    ResponseGroup='Large', Sort='salesrank')


                nodes = amazon.browse_node_lookup(ResponseGroup='TopSellers', BrowseNodeId='9103696011')
                for n in nodes:
                    print(n)


            else:
                products = amazon.search(
                    Keywords='{} shirt'.format(tag), SearchIndex='Apparel', BrowseNode="9103696011",
                    ResponseGroup='Large')

            count = 0



            for it in products.iterate_pages():

                time.sleep(1)

                for product in products:
                    count += 1
                    print("Getting product", count, product.asin, product.sales_rank)
                    nodes = []
                    product_type = 2

                    if not product.sales_rank:
                        continue

                    if range == 'low' and int(product.sales_rank) > 100000:
                        continue

                    if range == 'high' and product.sales_rank < 100000:
                        continue

                    for node in product.browse_nodes:
                        for n in node.ancestors:
                            nodes.append(n.name)
                        nodes.append(node.name)
                        if node.name == "Women" or "women" in product.title.lower() or "woman" in product.title.lower():
                            product_type = 1
                        elif node.name == "Men" or "men" in product.title.lower() or "man" in product.title.lower():
                            product_type = 0

                    if "Novelty" in nodes:
                        pass
                    else:
                        continue

                    data['result'].append({
                        'title': product.title,
                        'sales_rank': product.sales_rank,
                        'monthly_sales_estimate': monthly_sales_estimate(product.sales_rank),
                        'asin': product.asin,
                        'small_image_url': product.large_image_url,
                        'reviews': product.reviews,
                        'detail_page_url': product.detail_page_url,
                        'features': product.features,
                        'type': product_type
                    })

            amazon_products.after_response(data)

        return data

    def get(self, request, format=None):
        """
        Return amazon products
        """
        try:
            data = self.getProducts(request, format)


        except Exception as e:

            print("Exception occured", e)
            traceback.print_exc()
            time.sleep(2)
            data = self.getProducts(request, format)

            data = {'result': []}
            '''
            tags = request.GET.get('tags', '')
            for tag in tags.split(','):
                products = Product.objects.filter(title__icontains = tag)
                for product in products:
                    data['result'].append({
                        'title': product.title,
                        'sales_rank': product.sales_rank,
                        'monthly_sales_estimate': monthly_sales_estimate(product.sales_rank),
                        'asin': product.asin,
                        'small_image_url': product.small_image_url,
                        'reviews': product.reviews,
                        'detail_page_url': product.detail_page_url,
                        'features': product.features,
                        'type': product.type
                    })
            '''

        return Response(data)


class TrademarksView(LoggingMixin, GenericAPIView):
    serializer_class = TrademarkSerializer

    def post(self, request, format=None):
        """
        Return check trademarks
        """
        result = requests.post(
            'http://52.41.13.151/v1/api/search/', data=request.data)
        return Response(result.content)


class SynonymsView(LoggingMixin, GenericAPIView):
    authentication_classes = (
        BasicAuthentication, TokenAuthentication, SessionAuthentication)

    def get(self, request, format=None):
        """
        Return words with synonyms
        """
        user = request.user
        user.backend = 'django.contrib.auth.backends.ModelBackend'
        login(request, user)
        data = {}
        data['synonyms'] = []
        tags = request.GET.get('tags', '')
        for tag in tags.split(','):
            result = requests.get('https://wordsapiv1.p.mashape.com/words/{0}/synonyms'.format(tag),
                                  headers={
                "X-Mashape-Key": settings.MASHAPE,
                "Accept": "application/json",
            })
            if 'synonyms' in result.json():
                data['synonyms'] += result.json()['synonyms']
        return Response(data)


class AntonymsView(LoggingMixin, GenericAPIView):
    authentication_classes = (
        BasicAuthentication, TokenAuthentication, SessionAuthentication)

    def get(self, request, format=None):
        """
        Return words with antonyms
        """
        user = request.user
        user.backend = 'django.contrib.auth.backends.ModelBackend'
        login(request, user)
        data = {}
        data['antonyms'] = []
        tags = request.GET.get('tags', '')
        for tag in tags.split(','):
            result = requests.get('https://wordsapiv1.p.mashape.com/words/{0}/antonyms'.format(tag),
                                  headers={
                "X-Mashape-Key": settings.MASHAPE,
                "Accept": "application/json",
            })
            if 'antonyms' in result.json():
                data['antonyms'] += result.json()['antonyms']
        return Response(data)


class KeywordToolView(LoggingMixin, GenericAPIView):
    authentication_classes = (
        BasicAuthentication, TokenAuthentication, SessionAuthentication)

    def get(self, request, format=None):
        """
        Return result from keywordtool
        """
        user = request.user
        user.backend = 'django.contrib.auth.backends.ModelBackend'
        login(request, user)
        keywords = request.GET.get('tags', '')
        result = {
            'keywords': [],
            'trademark': 0
        }

        if keywords != '':
            for word in keywords.split(','):
                payload = {
                    'apikey': settings.KEYWORDTOOL,
                    'keyword': '[{0}]'.format(word),
                    'output': 'json',
                    'country': 'us',
                    'language': 'en',
                    'metrics': 'true',
                    'metrics_location': '2840',
                    'metrics_language': 'en'
                }
                word = word.lower()

                word_result = Word.objects.filter(name=word).first()

                if word_result:
                    data = word_result.results
                else:
                    data = False
                    try:
                        data_keywordtool = requests.get(
                            'http://api.keywordtool.io/v2/search/suggestions/amazon', params=payload)
                        if data_keywordtool.status_code == 200:
                            results = data_keywordtool.json()
                            created_word = Word.objects.create(
                                name=word, results=results)
                            data = created_word.results
                    except Exception as e:
                        pass

                list_keywords = []
                if data:
                    for item in data['results']:
                        for sub_item in data['results'][item]:
                            if 'volume' in sub_item and 'string' in sub_item:
                                if sub_item['volume'] > 300 and not sub_item['string'] in list_keywords:
                                    list_keywords.append(sub_item['string'])
                                    result['keywords'].append({'name': sub_item['string'].replace(
                                        '[', '').replace(']', ''), 'volume': sub_item['volume'], 'trademark': False})

        return Response(result)


class ExportTemplatesView(GenericAPIView):
    authentication_classes = (
        BasicAuthentication, TokenAuthentication, SessionAuthentication)
    serializer_class = ExportSerializer

    def post(self, request, format=None):
        """
        Return excel file
        """
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid():
            serializer.save()
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'TEE SHIRT DESIGN'
            if serializer.data['photo']:
                http = urllib3.PoolManager()
                r = http.request('GET', serializer.data['xls_photo'])
                image_file = io.BytesIO(r.data)
                img = Image(image_file)
                ws.add_image(img, 'A2')
            ws['B1'] = 'BRAND NAME'
            ws['B2'] = serializer.data['brand_name']
            ws['C1'] = 'PRODUCT TITLE'
            ws['C2'] = serializer.data['title']
            ws['D1'] = 'BULLET POINT ONE'
            ws['D2'] = serializer.data['tags']
            ws['E1'] = 'BULET POINT TWO'
            ws['E2'] = serializer.data['main_tags']
            ws['F1'] = 'DESCRIPTION BOX'
            ws['F2'] = serializer.data['description']
            ws['G1'] = 'SELECTED KEYWORDS FROM WC'
            ws['G2'] = serializer.data['keywords']

            ws.column_dimensions["A"].width = 60
            ws.column_dimensions["B"].width = 60
            ws.column_dimensions["C"].width = 60
            ws.column_dimensions["D"].width = 60
            ws.column_dimensions["E"].width = 60
            ws.column_dimensions["F"].width = 60
            ws.column_dimensions["G"].width = 60
            ws.row_dimensions[2].height = 100

            filename = int(time.time())
            s3c = boto3.client('s3',
                               aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                               aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY)
            handle_xls = StringIO(save_virtual_workbook(wb))
            xls_file = s3c.put_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                                      Key='exel/{0}.xlsx'.format(filename), Body=handle_xls.read())
            result = {
                'data': serializer.data,
                'file': '{0}exel/{1}.xlsx'.format(settings.AWS_S3_ROOT, filename)
            }
            return Response(result)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ExportKeywordsView(GenericAPIView):
    serializer_class = ExportSerializer

    def post(self, request, format=None):
        """
        Save all keywrods to xls
        """
        filename = int(time.time())
        s3c = boto3.client('s3',
                           aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                           aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY)
        handle_txt = StringIO(request.data.get('keywords', ''))
        txt_file = s3c.put_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                                  Key='txt/{0}.txt'.format(filename), Body=handle_txt.read())
        result = {
            'file': '{0}txt/{1}.txt'.format(settings.AWS_S3_ROOT, filename)
        }
        return Response(result)


class ShopList(LoggingMixin, GenericAPIView):
    serializer_class = ShopSerializer
    authentication_classes = (
        TokenAuthentication, BasicAuthentication,  SessionAuthentication)

    def get(self, request, format=None):
        """
        Return list of shops
        """
        user = request.user
        user.backend = 'django.contrib.auth.backends.ModelBackend'
        login(request, user)
        shops = Shop.objects.filter(disabled=False)
        serializer = self.serializer_class(
            shops, many=True, context={'request': request})
        return Response(serializer.data)

    def post(self, request, format=None):
        """
        Return new template
        """
        data = request.data.copy()
        data['user'] = request.user.id
        serializer = TemplateSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class SubscribeView(LoggingMixin, GenericAPIView):
    serializer_class = SubscribeSerializer

    def post(self, request, format=None):
        """
        Return add email to subscribe
        """
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class UserDetailsView(RetrieveUpdateAPIView):
    authentication_classes = (
        BasicAuthentication, TokenAuthentication, SessionAuthentication)

    """
    Reads and updates UserModel fields
    Accepts GET, PUT, PATCH methods.

    Default accepted fields: username, first_name, last_name
    Default display fields: pk, username, email, first_name, last_name
    Read-only fields: pk, email

    Returns UserModel fields.
    """
    serializer_class = UserDetailsSerializer

    def get_object(self):
        user = self.request.user
        user.backend = 'django.contrib.auth.backends.ModelBackend'
        login(self.request, user)
        return self.request.user

    def get_queryset(self):
        """
        Adding this method since it is sometimes called when using
        django-rest-swagger
        https://github.com/Tivix/django-rest-auth/issues/275
        """
        return get_user_model().objects.none()
